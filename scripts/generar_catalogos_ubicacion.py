from pathlib import Path
from urllib.request import urlretrieve
from openpyxl import load_workbook

# Fichero oficial del INE: relación de municipios y códigos por provincias.
INE_XLSX_URL = "https://www.ine.es/daco/daco42/codmun/diccionario26.xlsx"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_SQL = PROJECT_ROOT / "db" / "03_catalogos_ubicacion.sql"
TEMP_XLSX = PROJECT_ROOT / "scripts" / "diccionario26.xlsx"

PROVINCIAS = {
    "01": "Araba/Álava",
    "02": "Albacete",
    "03": "Alicante/Alacant",
    "04": "Almería",
    "05": "Ávila",
    "06": "Badajoz",
    "07": "Balears, Illes",
    "08": "Barcelona",
    "09": "Burgos",
    "10": "Cáceres",
    "11": "Cádiz",
    "12": "Castellón/Castelló",
    "13": "Ciudad Real",
    "14": "Córdoba",
    "15": "Coruña, A",
    "16": "Cuenca",
    "17": "Girona",
    "18": "Granada",
    "19": "Guadalajara",
    "20": "Gipuzkoa",
    "21": "Huelva",
    "22": "Huesca",
    "23": "Jaén",
    "24": "León",
    "25": "Lleida",
    "26": "Rioja, La",
    "27": "Lugo",
    "28": "Madrid",
    "29": "Málaga",
    "30": "Murcia",
    "31": "Navarra",
    "32": "Ourense",
    "33": "Asturias",
    "34": "Palencia",
    "35": "Palmas, Las",
    "36": "Pontevedra",
    "37": "Salamanca",
    "38": "Santa Cruz de Tenerife",
    "39": "Cantabria",
    "40": "Segovia",
    "41": "Sevilla",
    "42": "Soria",
    "43": "Tarragona",
    "44": "Teruel",
    "45": "Toledo",
    "46": "Valencia/València",
    "47": "Valladolid",
    "48": "Bizkaia",
    "49": "Zamora",
    "50": "Zaragoza",
    "51": "Ceuta",
    "52": "Melilla",
}


def sql_escape(value: str) -> str:
    return str(value).replace("\\", "\\\\").replace("'", "''")


def chunks(items, size):
    for i in range(0, len(items), size):
        yield items[i:i + size]


def main() -> None:
    print("Descargando fichero del INE...")
    urlretrieve(INE_XLSX_URL, TEMP_XLSX)

    print("Leyendo municipios...")
    workbook = load_workbook(TEMP_XLSX, read_only=True, data_only=True)
    sheet = workbook.active

    municipios = []

    # El Excel suele tener:
    # fila 1: título
    # fila 2: cabeceras CODAUTO, CPRO, CMUN, DC, NOMBRE
    # desde fila 3: datos
    for row in sheet.iter_rows(min_row=3, values_only=True):
        codauto, cpro, cmun, dc, nombre = row

        if cpro is None or cmun is None or nombre is None:
            continue

        cpro = str(cpro).zfill(2)
        cmun = str(cmun).zfill(3)
        codigo_ine = cpro + cmun
        nombre = str(nombre).strip()

        municipios.append((cpro, codigo_ine, nombre))

    print(f"Municipios encontrados: {len(municipios)}")

    OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)

    with OUTPUT_SQL.open("w", encoding="utf-8") as file:
        file.write("-- =========================================\n")
        file.write("-- 03_catalogos_ubicacion.sql\n")
        file.write("-- Catálogo de provincias y municipios de España\n")
        file.write("-- Fuente: INE - Relación de municipios y códigos por provincias\n")
        file.write("-- Generado automáticamente por scripts/generar_catalogos_ubicacion.py\n")
        file.write("-- =========================================\n\n")

        file.write("USE mascotas_webapp;\n\n")
        file.write("SET NAMES utf8mb4;\n\n")

        file.write("-- =========================================\n")
        file.write("-- PROVINCIAS\n")
        file.write("-- =========================================\n")
        file.write("INSERT INTO provincias (codigo_ine, nombre) VALUES\n")

        provincia_values = []
        for codigo, nombre in sorted(PROVINCIAS.items()):
            provincia_values.append(f"('{codigo}', '{sql_escape(nombre)}')")

        file.write(",\n".join(provincia_values))
        file.write("\nON DUPLICATE KEY UPDATE\n")
        file.write("    nombre = VALUES(nombre);\n\n")

        file.write("-- =========================================\n")
        file.write("-- MUNICIPIOS\n")
        file.write("-- =========================================\n")
        file.write("DROP TEMPORARY TABLE IF EXISTS tmp_municipios;\n\n")
        file.write("""
CREATE TEMPORARY TABLE tmp_municipios (
    cpro CHAR(2) NOT NULL,
    codigo_ine CHAR(5) NOT NULL,
    nombre VARCHAR(150) NOT NULL
) ENGINE=Memory;

""")

        for bloque in chunks(municipios, 1000):
            file.write("INSERT INTO tmp_municipios (cpro, codigo_ine, nombre) VALUES\n")
            values = []
            for cpro, codigo_ine, nombre in bloque:
                values.append(
                    f"('{cpro}', '{codigo_ine}', '{sql_escape(nombre)}')"
                )
            file.write(",\n".join(values))
            file.write(";\n\n")

        file.write("""
INSERT INTO municipios (provincia_id, codigo_ine, nombre)
SELECT
    p.id,
    t.codigo_ine,
    t.nombre
FROM tmp_municipios t
INNER JOIN provincias p ON p.codigo_ine = t.cpro
ON DUPLICATE KEY UPDATE
    provincia_id = VALUES(provincia_id),
    nombre = VALUES(nombre);

DROP TEMPORARY TABLE IF EXISTS tmp_municipios;
""")

    print(f"SQL generado correctamente en: {OUTPUT_SQL}")


if __name__ == "__main__":
    main()